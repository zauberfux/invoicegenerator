import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIG
# =============================================================================

DEFAULT_BUSINESS_FIELD = "00"   # user edits in Excel (B2)
HOURS_PER_DAY = 8
EUR_FORMAT = u"€#,##0.00"
USER_INPUT_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")  # light blue


# =============================================================================
# Helpers
# =============================================================================

@dataclass
class FileMeta:
    person_name: str
    start_date: Optional[datetime]
    end_date: Optional[datetime]
    # used for filename (keep legacy behavior: Month Year if same month)
    time_period_label: str

    @property
    def time_period_field(self) -> str:
        # shown in Excel cell B3: always exact from–to if parseable
        if not self.start_date or not self.end_date:
            return self.time_period_label
        return f"{self.start_date.strftime('%Y-%m-%d')} to {self.end_date.strftime('%Y-%m-%d')}"


def parse_filename_meta(filename: str) -> FileMeta:
    """
    Expected pattern:
      "<Name>-LoggedTime-YYYYMMDD-YYYYMMDD.csv"
    """
    stem = Path(filename).stem
    m = re.match(r"^(?P<person>.+?)-LoggedTime-(?P<start>\d{8})-(?P<end>\d{8})$", stem)
    if not m:
        return FileMeta(person_name=stem, start_date=None, end_date=None, time_period_label="Unknown period")

    person = m.group("person").strip()
    start_dt = datetime.strptime(m.group("start"), "%Y%m%d")
    end_dt = datetime.strptime(m.group("end"), "%Y%m%d")

    # keep filename behavior: if same month => "January 2026", else from-to
    if start_dt.year == end_dt.year and start_dt.month == end_dt.month:
        period_label = start_dt.strftime("%B %Y")
    else:
        period_label = f"{start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}"

    return FileMeta(person_name=person, start_date=start_dt, end_date=end_dt, time_period_label=period_label)


def excel_num_invariant(x: float) -> str:
    """XLSX formulas are locale-independent. Use dot decimals always."""
    return f"{x:.10f}".rstrip("0").rstrip(".")


def company_from_project_code_str(code_str: str) -> str:
    s = (code_str or "").strip()
    if s.startswith("1"):
        return "PCG"
    if s.startswith("2"):
        return "PCR"
    return "UNASSIGNED"


def is_missing_code(code_str: str) -> bool:
    s = (code_str or "").strip()
    if s == "":
        return True
    return not re.fullmatch(r"\d+", s)


def display_len_for_autosize(v) -> int:
    # prevent long formulas from making columns huge
    if v is None:
        return 0
    if isinstance(v, str) and v.startswith("="):
        return 10
    return len(str(v))


def autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        for cell in col:
            max_len = max(max_len, display_len_for_autosize(cell.value))
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(8, min(max_len + 2, 45))


def write_table(ws, table_rows: pd.DataFrame, start_row: int, title: str, day_rate_cell: str) -> Tuple[int, int]:
    ws[f"A{start_row}"] = title
    ws[f"A{start_row}"].font = Font(bold=True)

    ws.append(["Project Code", "Project", "Logged hrs", "Days", "Day Rate", "Costs"])
    for cell in ws[start_row + 1]:
        cell.font = Font(bold=True)

    start_row += 2
    data_start = start_row

    for _, r in table_rows.iterrows():
        ws.append([None, None, None, None, None, None])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=1).value = r["Project code"]
        ws.cell(row=row_idx, column=2).value = r["Project"]
        ws.cell(row=row_idx, column=3).value = r["Logged hrs"]

    data_end = data_start + len(table_rows) - 1

    for i in range(data_start, data_end + 1):
        ws[f"D{i}"] = f"=C{i}/{HOURS_PER_DAY}"
        ws[f"E{i}"] = f"={day_rate_cell}"
        ws[f"E{i}"].number_format = EUR_FORMAT
        ws[f"F{i}"] = f"=D{i}*E{i}"
        ws[f"F{i}"].number_format = EUR_FORMAT

    ws[f"E{data_end + 1}"] = "Subtotal:"
    ws[f"E{data_end + 1}"].font = Font(bold=True)
    ws[f"F{data_end + 1}"] = f"=SUM(F{data_start}:F{data_end})"
    ws[f"F{data_end + 1}"].font = Font(bold=True)
    ws[f"F{data_end + 1}"].number_format = EUR_FORMAT

    return data_end + 3, data_end + 1


def build_invoice_xlsx_bytes(df: pd.DataFrame, meta: FileMeta) -> BytesIO:
    required = {"Project", "Project code", "Logged Billable hours", "Logged Non-billable hours"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"CSV missing required columns: {sorted(missing)}")

    # normalize
    df["Project code str"] = df["Project code"].astype("string").fillna("").str.strip()
    df["Logged Billable hours"] = df["Logged Billable hours"].fillna(0).astype(float)
    df["Logged Non-billable hours"] = df["Logged Non-billable hours"].fillna(0).astype(float)

    total_billable_hrs = float(df["Logged Billable hours"].sum())
    total_nonbillable_hrs_logged = float(df["Logged Non-billable hours"].sum())
    total_logged_hrs = total_billable_hrs + total_nonbillable_hrs_logged

    df["code_missing"] = df["Project code str"].apply(is_missing_code)
    df["Company"] = df["Project code str"].apply(company_from_project_code_str)

    pcg_billable_hrs = float(df.loc[(df["Company"] == "PCG") & (~df["code_missing"]), "Logged Billable hours"].sum())
    pcr_billable_hrs = float(df.loc[(df["Company"] == "PCR") & (~df["code_missing"]), "Logged Billable hours"].sum())
    denom = pcg_billable_hrs + pcr_billable_hrs

    # billable-only project listing (avoid double counting)
    billable = (
        df.groupby(["Project code str", "Project"], as_index=False)["Logged Billable hours"]
        .sum()
        .rename(columns={"Logged Billable hours": "Logged hrs"})
        .sort_values(["Project code str", "Project"], kind="stable")
    )
    billable["code_missing"] = billable["Project code str"].apply(is_missing_code)
    billable["Company"] = billable["Project code str"].apply(company_from_project_code_str)

    pcg = billable[(billable["Company"] == "PCG") & (~billable["code_missing"]) & (billable["Logged hrs"] > 0)].copy()
    pcr = billable[(billable["Company"] == "PCR") & (~billable["code_missing"]) & (billable["Logged hrs"] > 0)].copy()
    other = billable[(billable["code_missing"]) & (billable["Logged hrs"] > 0)].copy()  # billable but no usable code

    pcg["Project code"] = pcg["Project code str"].astype(int)
    pcr["Project code"] = pcr["Project code str"].astype(int)
    other["Project code"] = ""

    pcg = pcg[["Project code", "Project", "Logged hrs"]]
    pcr = pcr[["Project code", "Project", "Logged hrs"]]
    other = other[["Project code", "Project", "Logged hrs"]]

    # BF General rows allocate: logged nonbillable + PTO(B6+B7+B8) split by billable share (or 50/50 if denom==0)
    PTO_SUM = "(B6+B7+B8)"
    lnb = excel_num_invariant(total_nonbillable_hrs_logged)
    pcg_b = excel_num_invariant(pcg_billable_hrs)
    pcr_b = excel_num_invariant(pcr_billable_hrs)

    if denom == 0:
        pcg_bf_logged_formula = f"=({lnb}+{PTO_SUM})*0.5"
        pcr_bf_logged_formula = f"=({lnb}+{PTO_SUM})*0.5"
    else:
        pcg_bf_logged_formula = f"=({lnb}+{PTO_SUM})*({pcg_b}/({pcg_b}+{pcr_b}))"
        pcr_bf_logged_formula = f"=({lnb}+{PTO_SUM})*({pcr_b}/({pcg_b}+{pcr_b}))"

    bf_pcg_row = {
        "Project code": '=VALUE("1"&$B$2&"000")',
        "Project": '="BF"&$B$2&" General (PCG)"',
        "Logged hrs": pcg_bf_logged_formula,
    }
    bf_pcr_row = {
        "Project code": '=VALUE("2"&$B$2&"000")',
        "Project": '="BF"&$B$2&" General (PCR)"',
        "Logged hrs": pcr_bf_logged_formula,
    }

    pcg = pd.concat([pcg, pd.DataFrame([bf_pcg_row])], ignore_index=True)
    pcr = pd.concat([pcr, pd.DataFrame([bf_pcr_row])], ignore_index=True)

    # sort BF General last
    def is_bf_general(v) -> int:
        return 1 if isinstance(v, str) and "BF" in v and "General" in v else 0

    pcg["_bf"] = pcg["Project"].apply(is_bf_general)
    pcr["_bf"] = pcr["Project"].apply(is_bf_general)
    pcg = pcg.sort_values(["_bf", "Project code", "Project"], kind="stable").drop(columns=["_bf"])
    pcr = pcr.sort_values(["_bf", "Project code", "Project"], kind="stable").drop(columns=["_bf"])

    # --- workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws["A1"] = "Name:"; ws["B1"] = meta.person_name
    ws["A2"] = "Business Field:"; ws["B2"] = DEFAULT_BUSINESS_FIELD
    ws["A3"] = "Time Period:"; ws["B3"] = meta.time_period_field

    ws["A4"] = "Monthly Salary:"; ws["B4"] = 0.0
    ws["B4"].number_format = EUR_FORMAT

    ws["A5"] = "Number of Days Worked:"
    ws["B5"] = f"=({excel_num_invariant(total_logged_hrs)})/{HOURS_PER_DAY}"

    ws["A6"] = "Paid vacation hrs:"; ws["B6"] = 0.0
    ws["A7"] = "Paid sick leave hrs:"; ws["B7"] = 0.0
    ws["A8"] = "Paid public holiday hrs:"; ws["B8"] = 0.0

    ws["A9"] = "Paid Time-off Days:"; ws["B9"] = f"=(B6+B7+B8)/{HOURS_PER_DAY}"
    ws["A10"] = "Total days:"; ws["B10"] = "=B5+B9"

    ws["A11"] = "Day Rate:"; ws["B11"] = "=B4/B10"
    ws["B11"].number_format = EUR_FORMAT

    # mark user-fill cells blue
    for addr in ("B2", "B4", "B6", "B7", "B8"):
        ws[addr].fill = USER_INPUT_FILL

    row = 14
    row, pcg_sub = write_table(ws, pcg, row, "PCG Projects", day_rate_cell="$B$11") if len(pcg) else (row, None)
    row, pcr_sub = write_table(ws, pcr, row, "PCR Projects", day_rate_cell="$B$11") if len(pcr) else (row, None)
    row, oth_sub = write_table(ws, other, row, "Other (billable, no project code)", day_rate_cell="$B$11") if len(other) else (row, None)

    subs = [x for x in [pcg_sub, pcr_sub, oth_sub] if x is not None]
    if subs:
        ws[f"E{row}"] = "Grand Total:"; ws[f"E{row}"].font = Font(bold=True)
        ws[f"F{row}"] = "=" + "+".join([f"F{s}" for s in subs])
        ws[f"F{row}"].number_format = EUR_FORMAT
        ws[f"F{row}"].font = Font(bold=True)

    autosize_columns(ws)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def make_output_filename(meta: FileMeta) -> str:
    # keep existing naming behavior (do NOT change the period style here)
    out_name_safe = re.sub(r"[^A-Za-z0-9_-]+", "_", meta.person_name).strip("_")
    out_period_safe = re.sub(r"[^A-Za-z0-9_-]+", "_", meta.time_period_label).strip("_")
    return f"Invoice_{out_name_safe}_{out_period_safe}.xlsx"


# =============================================================================
# Streamlit UI
# =============================================================================

st.set_page_config(page_title="Invoice Generator", layout="centered")
st.title("Invoice Generator")

st.write(
    """
Go to Float, select your name and the relevant time period, and download **"time tracking data"**.
Upload the CSV file here.
"""
)

uploaded = st.file_uploader("Upload CSV", type=["csv"])

if uploaded is not None:
    try:
        meta = parse_filename_meta(uploaded.name)
        df = pd.read_csv(uploaded)

        # Light validation early to give immediate feedback
        required = {"Project", "Project code", "Logged Billable hours", "Logged Non-billable hours"}
        missing = required - set(df.columns)
        if missing:
            st.error(f"CSV is missing required columns: {sorted(missing)}")
            st.caption("Columns found:")
            st.code(", ".join(df.columns.astype(str).tolist()))
        else:
            st.success(f"Loaded: {meta.person_name} — {meta.time_period_field}")

            with st.form("generate_form"):
                generate = st.form_submit_button("Generate")

            if generate:
                xlsx_bytes = build_invoice_xlsx_bytes(df, meta)
                st.download_button(
                    label="Download Invoice",
                    data=xlsx_bytes,
                    file_name=make_output_filename(meta),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                st.info(
                    """
In the Excel file, fill in the values in the blue cells.

You find your time off hrs on your reports page, under **"Time off"**.
Please don’t count **"Ausgleich für zusätzliche Arbeitszeit"**, as this is already accounted for in previous months.
"""
                )

    except Exception as e:
        st.error("Could not process this file.")
        st.exception(e)
